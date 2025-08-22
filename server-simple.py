from flask import Flask, jsonify, request, send_file, send_from_directory
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import uuid
from datetime import datetime
import io
import os
import json

app = Flask(__name__)
app.config['SECRET_KEY'] = 'quiz-admin-secret-key'
CORS(app)

# Uploads directory for quiz JSON
UPLOADS_DIR = 'uploads'
os.makedirs(UPLOADS_DIR, exist_ok=True)
QUIZ_JSON_PATH = os.path.join(UPLOADS_DIR, 'quiz.json')
SAMPLE_QUIZ_PATH = os.path.join(os.path.dirname(__file__), 'sample-quiz.json')

# Quiz data is loaded from disk or upload (no hardcoded fallback)
quiz_data = None

# Store data
connected_clients = {}
quiz_results = []
quiz_in_progress = False
start_timestamp = None
current_question_index = 0


def validate_quiz_payload(payload: dict):
    """Validate minimal quiz schema."""
    if not isinstance(payload, dict):
        return False, 'Payload must be a JSON object'
    required_top = ["title", "timeLimit", "questions"]
    for key in required_top:
        if key not in payload:
            return False, f'Missing field: {key}'
    if not isinstance(payload["title"], str):
        return False, 'title must be a string'
    if not isinstance(payload["timeLimit"], int) or payload["timeLimit"] <= 0:
        return False, 'timeLimit must be a positive integer (seconds)'
    if not isinstance(payload["questions"], list) or len(payload["questions"]) == 0:
        return False, 'questions must be a non-empty array'
    for idx, q in enumerate(payload["questions"], start=1):
        if not isinstance(q, dict):
            return False, f'Question {idx} must be an object'
        for rq in ["id", "question", "options", "correctAnswer", "points"]:
            if rq not in q:
                return False, f'Question {idx}: missing field {rq}'
        if not isinstance(q["id"], int):
            return False, f'Question {idx}: id must be an integer'
        if not isinstance(q["question"], str) or not q["question"].strip():
            return False, f'Question {idx}: question must be a non-empty string'
        if not isinstance(q["options"], list) or len(q["options"]) < 2:
            return False, f'Question {idx}: options must be an array with at least 2 items'
        if not isinstance(q["correctAnswer"], int) or not (0 <= q["correctAnswer"] < len(q["options"])):
            return False, f'Question {idx}: correctAnswer must be a valid option index'
        if not isinstance(q["points"], int) or q["points"] <= 0:
            return False, f'Question {idx}: points must be a positive integer'
    return True, None


def load_quiz_from_disk():
    """Load quiz JSON from disk if present; update global quiz_data."""
    global quiz_data
    # Prefer uploaded quiz if present
    if os.path.exists(QUIZ_JSON_PATH):
        try:
            with open(QUIZ_JSON_PATH, 'r', encoding='utf-8') as f:
                payload = json.load(f)
            ok, err = validate_quiz_payload(payload)
            if not ok:
                print(f"Invalid quiz.json schema: {err}. No quiz loaded.")
                quiz_data = None
                return
            quiz_data = payload
            print(f"Loaded quiz from {QUIZ_JSON_PATH}: {quiz_data.get('title')} ({len(quiz_data.get('questions', []))} questions)")
            return
        except Exception as e:
            print(f"Failed to load quiz.json: {e}. No quiz loaded.")

    # Fallback to bundled sample-quiz.json for convenience
    if os.path.exists(SAMPLE_QUIZ_PATH):
        try:
            with open(SAMPLE_QUIZ_PATH, 'r', encoding='utf-8') as f:
                payload = json.load(f)
            ok, err = validate_quiz_payload(payload)
            if not ok:
                print(f"Invalid sample-quiz.json schema: {err}. No quiz loaded.")
                quiz_data = None
                return
            quiz_data = payload
            print(f"Loaded sample quiz: {quiz_data.get('title')} ({len(quiz_data.get('questions', []))} questions)")
        except Exception as e:
            print(f"Failed to load sample-quiz.json: {e}. No quiz loaded.")
    else:
        print("No quiz.json found. Upload a quiz via the admin panel.")


# Attempt to load quiz from uploads on startup
load_quiz_from_disk()

# Serve static files
@app.route('/')
def admin_panel():
    """Serve the admin panel"""
    return send_from_directory('admin-panel', 'index.html')

@app.route('/client')
def client_system():
    """Serve the client system"""
    return send_from_directory('client-system', 'login.html')

@app.route('/admin-panel/<path:filename>')
def admin_static(filename):
    """Serve admin panel static files"""
    return send_from_directory('admin-panel', filename)

@app.route('/client-system/<path:filename>')
def client_static(filename):
    """Serve client system static files"""
    return send_from_directory('client-system', filename)

# Additional routes for client files
@app.route('/styles.css')
def client_styles():
    """Serve client styles.css"""
    return send_from_directory('client-system', 'styles.css')

@app.route('/client.js')
def client_js():
    """Serve client.js"""
    return send_from_directory('client-system', 'client.js')

@app.route('/quiz.js')
def quiz_js():
    """Serve quiz.js"""
    return send_from_directory('client-system', 'quiz.js')

@app.route('/api/status')
def get_status():
    return jsonify({
        'status': 'running',
        'quizInProgress': quiz_in_progress,
        'totalClients': len(connected_clients),
        'completedClients': len(quiz_results),
        'quizLoaded': quiz_data is not None,
        'quizTitle': quiz_data.get('title') if quiz_data else None,
        'numQuestions': len(quiz_data.get('questions', [])) if quiz_data else 0,
        'startAt': start_timestamp,
        'currentQuestionIndex': current_question_index
    })

@app.route('/api/quiz/start', methods=['POST'])
def start_quiz():
    global quiz_in_progress, start_timestamp, current_question_index
    if not quiz_data or not quiz_data.get('questions'):
        return jsonify({'error': 'No quiz loaded. Upload a quiz JSON first.'}), 400
    if quiz_in_progress:
        return jsonify({'error': 'Quiz already in progress'}), 400
    
    quiz_in_progress = True
    start_timestamp = datetime.utcnow().isoformat() + 'Z'
    current_question_index = 0
    for client in connected_clients.values():
        client['status'] = 'quiz-active'
    
    return jsonify({'message': 'Quiz started successfully', 'startAt': start_timestamp})

@app.route('/api/quiz/next', methods=['POST'])
def next_question():
    global current_question_index
    if not quiz_in_progress:
        return jsonify({'error': 'Quiz not in progress'}), 400
    if not quiz_data or not quiz_data.get('questions'):
        return jsonify({'error': 'No quiz loaded'}), 400
    total = len(quiz_data.get('questions', []))
    if current_question_index < total - 1:
        current_question_index += 1
        return jsonify({'message': 'Advanced to next question', 'currentQuestionIndex': current_question_index})
    else:
        return jsonify({'message': 'Already at last question', 'currentQuestionIndex': current_question_index})

@app.route('/api/quiz/reset', methods=['POST'])
def reset_quiz():
    global quiz_in_progress, quiz_results, start_timestamp, connected_clients, current_question_index
    quiz_in_progress = False
    start_timestamp = None
    quiz_results.clear()
    # Clear all connected clients so counts/lists reset fully
    connected_clients.clear()
    current_question_index = 0
    
    return jsonify({'message': 'Quiz reset successfully'})

@app.route('/api/quiz/upload', methods=['POST'])
def upload_quiz():
    """Upload a quiz JSON file; replace in-memory quiz; store to disk."""
    global quiz_data, quiz_in_progress, quiz_results
    if quiz_in_progress:
        return jsonify({'error': 'Cannot upload while quiz is in progress. Reset the quiz first.'}), 400

    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400

    try:
        content = file.read().decode('utf-8')
        payload = json.loads(content)
    except Exception as e:
        return jsonify({'error': f'Invalid JSON file: {e}'}), 400

    ok, err = validate_quiz_payload(payload)
    if not ok:
        return jsonify({'error': f'Invalid schema: {err}'}), 400

    # Persist to disk
    try:
        with open(QUIZ_JSON_PATH, 'w', encoding='utf-8') as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
    except Exception as e:
        return jsonify({'error': f'Failed to save file: {e}'}), 500

    # Replace in-memory quiz and reset states
    quiz_data = payload
    quiz_results.clear()
    for client in connected_clients.values():
        client['status'] = 'waiting'

    return jsonify({
        'message': 'Quiz uploaded successfully',
        'quizTitle': quiz_data.get('title'),
        'numQuestions': len(quiz_data.get('questions', []))
    })

@app.route('/api/quiz', methods=['GET'])
def get_quiz():
    if not quiz_data:
        return jsonify({'error': 'No quiz loaded'}), 404
    return jsonify(quiz_data)

def validate_quiz_result(result):
    """Validate and clean up quiz result data structure."""
    if not isinstance(result, dict):
        return False
    
    # Check if it's the old format and convert it
    if 'score' in result and 'maxScore' in result:
        # Old format - convert to new format
        try:
            score = result.get('score', 0)
            max_score = result.get('maxScore', 0)
            if max_score > 0:
                correct_answers = round((score / max_score) * 20)  # Convert to 20 questions
                total_questions = 20
                incorrect_answers = total_questions - correct_answers
                accuracy = round((correct_answers / total_questions) * 100)
            else:
                correct_answers = 0
                total_questions = 20
                incorrect_answers = 20
                accuracy = 0
            
            # Update the result with new structure
            result['correctAnswers'] = correct_answers
            result['totalQuestions'] = total_questions
            result['incorrectAnswers'] = incorrect_answers
            result['accuracy'] = accuracy
            
            # Remove old fields
            result.pop('score', None)
            result.pop('maxScore', None)
            result.pop('percentage', None)
            
            return True
        except Exception as e:
            print(f"Error converting old result format: {e}")
            return False
    
    # Check if it has the new format
    required_fields = ['correctAnswers', 'totalQuestions', 'incorrectAnswers', 'accuracy']
    return all(field in result for field in required_fields)

@app.route('/api/results')
def get_results():
    # Validate and clean up all results
    valid_results = []
    for result in quiz_results:
        if validate_quiz_result(result):
            valid_results.append(result)
        else:
            print(f"Invalid result removed: {result}")
    
    # Update the global quiz_results with cleaned data
    quiz_results.clear()
    quiz_results.extend(valid_results)
    
    sorted_results = sorted(quiz_results, key=lambda x: (-x['correctAnswers'], x.get('timeTaken', 0)))
    return jsonify({'results': sorted_results})

@app.route('/api/results/download')
def download_results():
    try:
        if not quiz_results:
            return jsonify({'error': 'No results available to download'}), 404
        
        # Validate all results before processing
        valid_results = []
        for result in quiz_results:
            if validate_quiz_result(result):
                valid_results.append(result)
            else:
                print(f"Invalid result found during download: {result}")
        
        if not valid_results:
            return jsonify({'error': 'No valid results available to download'}), 404
            
        wb = Workbook()
        ws = wb.active
        ws.title = "Quiz Results"
        
        # Sort results by correct answers (descending) and then by time taken
        sorted_results = sorted(valid_results, key=lambda x: (-x['correctAnswers'], x.get('timeTaken', 0)))
        
        # Add headers
        ws.cell(row=1, column=1, value='Rank')
        ws.cell(row=1, column=2, value='Player Name')
        ws.cell(row=1, column=3, value='Correct Answers')
        ws.cell(row=1, column=4, value='Total Questions')
        ws.cell(row=1, column=5, value='Incorrect Answers')
        ws.cell(row=1, column=6, value='Accuracy')
        
        # Add data
        for row, result in enumerate(sorted_results, start=2):
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value=result['clientName'])
            ws.cell(row=row, column=3, value=result['correctAnswers'])
            ws.cell(row=row, column=4, value=result['totalQuestions'])
            ws.cell(row=row, column=5, value=result['incorrectAnswers'])
            ws.cell(row=row, column=6, value=f"{result['accuracy']}%")
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='quiz-results.xlsx'
        )
    except Exception as e:
        print(f"Error downloading results: {e}")
        return jsonify({'error': f'Failed to generate Excel file: {str(e)}'}), 500

@app.route('/api/clients')
def get_clients():
    return jsonify({
        'totalClients': len(connected_clients),
        'clients': list(connected_clients.values())
    })

@app.route('/api/client/connect', methods=['POST'])
def client_connect():
    client_id = str(uuid.uuid4())
    data = request.get_json(silent=True) or {}
    provided_name = data.get('name')
    safe_name = None
    if isinstance(provided_name, str):
        stripped = provided_name.strip()
        if stripped:
            safe_name = stripped
    client_info = {
        'id': client_id,
        'name': safe_name or f'Client {len(connected_clients) + 1}',
        'status': 'quiz-active' if quiz_in_progress else 'waiting'
    }
    connected_clients[client_id] = client_info
    return jsonify({
        'clientId': client_id,
        'quizData': quiz_data,  # may be null if not uploaded yet
        'totalClients': len(connected_clients),
        'quizInProgress': quiz_in_progress,
        'startAt': start_timestamp,
        'currentQuestionIndex': current_question_index
    })

@app.route('/api/client/submit', methods=['POST'])
def submit_quiz():
    if not quiz_data or not quiz_data.get('questions'):
        return jsonify({'error': 'No quiz loaded'}), 400

    data = request.get_json()
    client_id = data.get('clientId')
    
    if client_id not in connected_clients:
        return jsonify({'error': 'Client not found'}), 404
    
    client = connected_clients[client_id]
    client['status'] = 'completed'
    
    # Debug logging
    print(f"Client {client['name']} submitted answers: {data.get('answers')}")
    print(f"Quiz has {len(quiz_data['questions'])} questions")
    
    correct_answers = 0
    total_questions = len(quiz_data['questions'])
    answers = data.get('answers', [])
    
    for i, answer in enumerate(answers):
        if i < total_questions:
            question = quiz_data['questions'][i]
            # Handle both string and numeric answers
            if isinstance(answer, str):
                try:
                    answer = int(answer)
                except ValueError:
                    answer = None
            
            if answer is not None and answer == question['correctAnswer']:
                correct_answers += 1
                print(f"Question {i+1}: Correct! Answer: {answer}")
            else:
                print(f"Question {i+1}: Wrong. Expected {question['correctAnswer']}, got {answer}")
    
    # Calculate based on question count, not points
    incorrect_answers = total_questions - correct_answers
    accuracy = round((correct_answers / total_questions) * 100) if total_questions > 0 else 0
    
    print(f"Final result: {correct_answers}/{total_questions} correct ({accuracy}%)")
    
    quiz_result = {
        'clientId': client_id,
        'clientName': client['name'],
        'correctAnswers': correct_answers,
        'totalQuestions': total_questions,
        'incorrectAnswers': incorrect_answers,
        'accuracy': accuracy,
        'timeTaken': 0
    }
    
    quiz_results.append(quiz_result)
    
    return jsonify({
        'message': 'Quiz submitted successfully',
        'correctAnswers': correct_answers,
        'totalQuestions': total_questions,
        'incorrectAnswers': incorrect_answers,
        'accuracy': accuracy
    })

if __name__ == '__main__':
    print("Starting Quiz Admin Panel Server (Simple Version)...")
    print("Server will run on: http://localhost:5000")
    print("Admin panel: http://localhost:5000/")
    print("Client system: http://localhost:5000/client")
    
    app.run(host='0.0.0.0', port=5000, debug=True)
